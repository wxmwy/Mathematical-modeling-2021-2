import numpy as np
import math
from openpyxl import load_workbook


def read_data(path):

    workbook = load_workbook(path)
    booksheet = workbook.active
    rows = booksheet.rows
    columns = booksheet.columns

    food = np.zeros((600,4))
    fire = np.zeros((600,4))

    start = np.zeros((3))
    end = np.zeros((3))

    fire_cnt = 0
    food_cnt = 0
    cnt = 0

    # 迭代所有的行
    for row in rows:
        cnt = cnt + 1
        cell_data_1 = booksheet.cell(row=cnt, column=2).value
        cell_data_2 = booksheet.cell(row=cnt, column=3).value
        cell_data_3 = booksheet.cell(row=cnt, column=4).value
        cell_data_4 = booksheet.cell(row=cnt, column=5).value
        cell_data_5 = booksheet.cell(row=cnt, column=6).value

        if (cell_data_5 == 0):
            if(cnt == 2):
                start[0] = cell_data_1
                start[1] = cell_data_2
                start[2] = cell_data_3
            else:
                end[0] = cell_data_1
                end[1] = cell_data_2
                end[2] = cell_data_3

        if(cell_data_5 != 0 and cnt != 1):
            if(cell_data_4 == 0):
                fire[fire_cnt][0] = cell_data_1
                fire[fire_cnt][1] = cell_data_2
                fire[fire_cnt][2] = cell_data_3
                fire[fire_cnt][3] = cell_data_5
                fire_cnt += 1
            else:
                food[food_cnt][0] = cell_data_1
                food[food_cnt][1] = cell_data_2
                food[food_cnt][2] = cell_data_3
                food[food_cnt][3] = cell_data_5
                food_cnt += 1

    fire.resize((fire_cnt, 4))
    food.resize((food_cnt, 4))

    return start, end, fire, food


# 求两点间距离
def point_dis(start, end):
    return math.sqrt((end[2] - start[2]) ** 2 + (end[1] - start[1]) ** 2 + (end[0] - start[0]) ** 2)


# 求两点间行走耗费的舒适/饱食值
def point_cost(start, end):
    if(start[2] == end[2]):
        return point_dis(start, end) / 100 * 5
    if (start[2] > end[2]):
        return point_dis(start, end) / 100 * 4
    return point_dis(start, end) / 100 * 6


# 求两向量夹角的cos值
def cos_vector(v1, v2):
    cos = (v1[0] * v2[0] + v1[1] * v2[1] + v1[2] * v2[2]) / (math.sqrt(v1[0] ** 2 + v1[1] ** 2 + v1[2] ** 2) * math.sqrt(v2[0] ** 2 + v2[1] ** 2 + v2[2] ** 2))
    #print(v1, v2, cos)
    return cos

def problem1(start, end, fire, food, opt):
    num = 0
    '''
    dis_now = point_dis(start, end)
    fire_earn = 0
    food_earn = 0
    fire_safe_point = 0
    fire_danger_point = 0
    food_safe_point = 0
    food_danger_point = 0
    for i in 100:
        min_state = min(opt.fire_state, opt.food_state)
        if(point_cost(opt.point_now, end) < min_state - opt.danger_line):
            print("find way out!!!")
            break
        for j in fire.shape[0]:
            cost = point_cost(opt.point_now, fire[j])
            if(cost < min_state - opt.safe_line):
                earn = fire[j][3] * (point_dis(opt.point_now, end) - point_dis(fire[j] - end))
                if(earn > fire_earn):
                    fire_earn = earn
                    fire_safe_point = j
        for k in food.shape[0]:
            cost = point_cost(opt.point_now, food[k])
            if (cost < min_state - opt.safe_line):
                earn = food[j][3] * (point_dis(opt.point_now, end) - point_dis(food[j] - end))
                if (earn > food_earn):
                    food_earn = earn
                    food_safe_point = j
        if(fire_earn == 0 and food_earn == 0):
            print("search path fail!")
            exit(-1)

        if(opt.fire_state > opt.food_state):
    '''

    print()
    return num


def problem2(start, end, fire, food, opt):
    dis = 0
    opt.point_now = start
    '''
    test_line = 10
    update_line = 0
    for i in range(100):
        print("@@@@start iter ", i)
        fire_earn = 0
        food_earn = 0
        fire_point = -1
        food_point = -1
        min_state = min(opt.fire_state, opt.food_state)
        v_now = end - opt.point_now
        if (point_cost(opt.point_now, end) < min_state - opt.danger_line):
            dis += point_dis(opt.point_now, end)
            print("find way out!!!")
            break

        for j in range(fire.shape[0]):
            if(opt.point_now[0] == fire[j][0] and opt.point_now[1] == fire[j][1] and opt.point_now[2] == fire[j][2]):
                continue
            cost = point_cost(opt.point_now, fire[j])
            if (cost < min_state - test_line):
                earn = cos_vector(fire[j][0:3] - opt.point_now, v_now)
                if (earn > fire_earn):
                    fire_earn = earn
                    fire_point = j
        for k in range(food.shape[0]):
            if (opt.point_now[0] == food[k][0] and opt.point_now[1] == food[k][1] and opt.point_now[2] == food[k][2]):
                continue
            cost = point_cost(opt.point_now, food[k])
            if (cost < min_state - test_line):
                earn = cos_vector(food[k][0:3] - opt.point_now, v_now)
                if (earn > food_earn):
                    food_earn = earn
                    food_point = k
        if(food_earn == 0 and fire_earn == 0):
            if(test_line <= -5):
                print("You dead in travel!")
                exit(-1)
            else:
                test_line = test_line - 1
                continue
        test_line = int(min_state - 1)
        if((opt.fire_state < opt.food_state and fire_earn != 0) or food_earn == 0):
            dis += point_dis(opt.point_now, fire[fire_point])
            cost = point_cost(opt.point_now, fire[fire_point])
            opt.fire_state = opt.fire_state - cost + fire[fire_point][3]
            opt.food_state = opt.food_state - cost
            opt.point_now = fire[fire_point][0:3]
        else:
            dis += point_dis(opt.point_now, food[food_point])
            cost = point_cost(opt.point_now, food[food_point])
            opt.food_state = opt.food_state - cost + food[food_point][3]
            opt.fire_state = opt.fire_state - cost
            opt.point_now = food[food_point][0:3]
        #print("i j ", i, " ", j)
        print("point next ", opt.point_now, " fire left ", opt.fire_state, " food left ", opt.food_state)
    '''
    return dis


def problem3(start, end, fire, food, opt):
    print()


class people_state():
    def __init__(self):
        self.fire_state = 10
        self.food_state = 10
        self.dead_line = -5
        self.danger_line = -3
        self.safe_line = 0
        self.point_now = []


def main():
    data_path = "C:\\Users\\lenovo\\Desktop\\path.xlsx"
    start, end, fire, food = read_data(data_path)
    print("fire_cnt is ", fire.shape)
    print("food_cnt is ", food.shape)
    print("start at ", start)
    print("end at ", end)
    opt1 = people_state()
    #problem1(start, end, fire, food, opt1)
    opt2 = people_state()
    problem2(start, end, fire, food, opt2)
    opt3 = people_state()
    #problem3(start, end, fire, food, opt3)


main()